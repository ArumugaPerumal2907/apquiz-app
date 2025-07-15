from flask import Flask
from flask import render_template, request, redirect, url_for, session, flash, jsonify
from werkzeug.utils import secure_filename
import mysql.connector
import bcrypt
import os
from openpyxl import load_workbook
from config import Config

app = Flask(__name__)
app.config.from_object(Config)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def get_db_connection():
    return mysql.connector.connect(
        host=app.config['MYSQL_HOST'],
        user=app.config['MYSQL_USER'],
        password=app.config['MYSQL_PASSWORD'],
        database=app.config['MYSQL_DB']
    )

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/')
def index():
    return render_template('index.html')

# Registration
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute('INSERT INTO users (email, password) VALUES (%s, %s)', (email, hashed))
            conn.commit()
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))
        except mysql.connector.Error:
            flash('Email already registered.', 'danger')
        finally:
            cur.close()
            conn.close()
    return render_template('register.html')

# Login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute('SELECT * FROM users WHERE email = %s', (email,))
        user = cur.fetchone()
        cur.close()
        conn.close()
        if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
            session['user_id'] = user['id']
            session['email'] = user['email']
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard.html')

# Excel Upload
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        file = request.files['file']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            ingest_excel(filepath)
            flash('Questions uploaded successfully!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid file type.', 'danger')
    return render_template('upload.html')

def ingest_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    conn = get_db_connection()
    cur = conn.cursor()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        question, a, b, c, d, correct = row
        cur.execute('''INSERT INTO questions (question, option_a, option_b, option_c, option_d, correct_ans) VALUES (%s, %s, %s, %s, %s, %s)''',
                    (question, a, b, c, d, correct))
    conn.commit()
    cur.close()
    conn.close()

# Quiz Start
@app.route('/quiz')
def quiz():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('SELECT * FROM questions')
    questions = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('quiz.html', questions=questions)

# Quiz Submission
@app.route('/submit', methods=['POST'])
def submit():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    answers = request.json.get('answers', {})
    conn = get_db_connection()
    cur = conn.cursor()
    score = 0
    results = []
    for qid, selected in answers.items():
        cur.execute('SELECT correct_ans FROM questions WHERE id = %s', (qid,))
        correct = cur.fetchone()[0]
        is_correct = (selected == correct)
        if is_correct:
            score += 1
        results.append({'question_id': qid, 'selected': selected, 'correct': correct, 'is_correct': is_correct})
        cur.execute('''INSERT INTO user_answers (user_id, question_id, selected_ans, is_correct) VALUES (%s, %s, %s, %s)''',
                    (session['user_id'], qid, selected, is_correct))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'score': score, 'results': results})

# Results Page
@app.route('/results')
def results():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute('''SELECT q.question, q.correct_ans, ua.selected_ans, ua.is_correct FROM user_answers ua JOIN questions q ON ua.question_id = q.id WHERE ua.user_id = %s''', (session['user_id'],))
    results = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('results.html', results=results)

if __name__ == '__main__':
    app.secret_key = app.config['SECRET_KEY']
    app.run(debug=True)
